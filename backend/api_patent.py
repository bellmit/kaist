# -*- coding: utf-8 -*-
print ("module [backend.api_patent] loaded")

from backend import manager
from backend.api_common import *
from backend_model.table_patent import *
from sqlalchemy import or_, and_,sql
import requests
import re
import os
#from pymongo import MongoClient
from datetime import datetime
from bs4 import BeautifulSoup
from konlpy.tag import Okt,Hannanum
# import pandas as pd
# import subprocess
from tqdm import tqdm
from tika import parser
# import numpy as np
db = DBManager.db
import json
from time import sleep
import uuid
import sys
if sys.version_info[0] < 3:
    import urllib
else :
    import urllib.parse
from urllib.parse import urlparse
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib

from konlpy.tag import Okt
import pandas as pd
from .modeling import *
from .data import *
import csv
from sqlalchemy import or_, and_, func
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment

#ELASTIC_SERVER = 'http://localhost:9200'
ELASTIC_SERVER = 'http://211.41.73.43:9200'
os.environ['CUDA_VISIBLE_DEVICES'] = '0'

api_headers = {'Content-type': 'application/json'}
#
# global g_Moodel
# WEIGHT_FILE_NAME = './models/cedrknrm_8000/weights.p'
# print("WEIGHT_FILE_NAME :",WEIGHT_FILE_NAME)
# g_Moodel = CedrKnrmRanker().cuda()
# g_Moodel.load(WEIGHT_FILE_NAME)

tech_types = {'기계/소재':0,'바이오/의료':1,'에너지/자원':2,'전기/전자':3,'정보통신/지식서비스':4,'화학':5,'기타':9}
tech_types_dict= {0:'기계/소재',1:'바이오/의료',2:'에너지/자원',3:'전기/전자',4:'정보통신/지식서비스',5:'화학',9:'기타'}

# from openpyxl import load_workbook
# from openpyxl.styles.borders import Border, Side, BORDER_THIN
# from openpyxl.styles import Alignment
def post_get_many_source(result=None, **kw):
    check_token()

    src_list = result['objects']

    for src in src_list:
        if src['invention_title'] is not None:
            src['invention_title'] = src['invention_title'][:100] + "..."

        if src['summary_of_invention'] is not None:
            src['summary_of_invention'] = src['summary_of_invention'][:300] + "..."

        if src['description_of_embodiments'] is not None:
            src['description_of_embodiments'] = src['description_of_embodiments'][:1000] + "..."

def post_get_many_train_result(result=None, **kw):
    check_token()
    src_list = result['objects']


manager.create_api(KitechReport
                   , url_prefix='/api/v1'
                   , collection_name='reports'
                   , methods=['GET', 'PUT', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [check_token],
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token_single],
                        'GET_MANY': [check_token]
                   })


def post_patch_single_train_result(result=None, **kw):
    check_token_single()
    print(result)
    KitechTrainResult.query.filter(KitechTrainResult.id != result['id']).update({'active': 0})
    db.session.commit()



manager.create_api(KitechTrainResult
                   , url_prefix='/api/v1'
                   , collection_name='trainresult'
                   , methods=['GET', 'PUT', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [check_token],
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token_single],
                        'GET_MANY': [check_token]
                   }
                   , postprocessors={
                        'PATCH_SINGLE': [post_patch_single_train_result]
                   })

@app.route('/api/v1/SearchQuery', methods=['GET'])
def searchquery_api():
    print("searchquery_api start")
    try :
        limit = request.args.get('limit')
        offset = request.args.get('offset')
        querystring = request.args.get('querystring')
        user_id = request.args.get('user_id')
        graph = request.args.get('graph')
        print("limit : ", limit, ", offset : ", offset, ", querystring : ", querystring, ",user_id : ",user_id)
        send_url = app.config["SEARCH_API_URL"] + "/api/SearchQuery?limit={0}&offset={1}&querystring={2}&user_id={3}&graph={4}".format(limit, offset,querystring,user_id,graph)
        print("send_url : ",send_url)
        response = requests.get(send_url)
        print("response :",response.text)
        # if int(graph) == 2 :
        #     saved_xls_file = app.config["STATIC_FILE_PATH"] + '/templete/score.xlsx'
        #     new_xls_file = app.config["STATIC_FILE_PATH"] + '/excel/score_maked.xlsx'
        #     wb = load_workbook(saved_xls_file, data_only=True)
        #     sheet = wb.active
        #     rows = sheet.rows
        #     cur_row = 2
        #     index = 1
        #     thin_border = Border(
        #         left=Side(border_style=BORDER_THIN, color='00000000'),
        #         right=Side(border_style=BORDER_THIN, color='00000000'),
        #         top=Side(border_style=BORDER_THIN, color='00000000'),
        #         bottom=Side(border_style=BORDER_THIN, color='00000000')
        #     )
        #     alignment = Alignment(horizontal='center',
        #                           vertical='bottom',
        #                           text_rotation=0,
        #                           wrap_text=False,
        #                           shrink_to_fit=True,
        #                           indent=0)
        #     data_objects = json.loads(response.text)
        #     for data in data_objects['meta']['ai_score']:
        #         print("data ai: ",data)
        #         sheet.cell(row=cur_row, column=1).value = str(index)
        #         sheet.cell(row=cur_row, column=2).value = data
        #         sheet.cell(row=cur_row, column=1).border = thin_border
        #         sheet.cell(row=cur_row, column=2).border = thin_border
        #         sheet.cell(row=cur_row, column=1).border = thin_border
        #         sheet.cell(row=cur_row, column=2).border = thin_border
        #         cur_row += 1
        #         index += 1
        #     cur_row = 2
        #     for data in data_objects['meta']['bm_score']:
        #         print("data bm: ",data)
        #         sheet.cell(row=cur_row, column=3).value = data
        #         sheet.cell(row=cur_row, column=3).border = thin_border
        #         cur_row += 1
        #     cur_row = 2
        #     for data in data_objects['meta']['patent_code']:
        #         print("data patent_code: ",data)
        #         sheet.cell(row=cur_row, column=4).value = data
        #         sheet.cell(row=cur_row, column=4).border = thin_border
        #         cur_row += 1
        #     wb.save(filename=new_xls_file)
        # if int(graph) == 3 :
        #     saved_xls_file = app.config["STATIC_FILE_PATH"] + '/templete/search.xlsx'
        #     new_xls_file = app.config["STATIC_FILE_PATH"] + '/excel/search_result.xlsx'
        #     wb = load_workbook(saved_xls_file, data_only=True)
        #     sheet = wb.active
        #     rows = sheet.rows
        #     cur_row = 4
        #     index = 1
        #     thin_border = Border(
        #         left=Side(border_style=BORDER_THIN, color='00000000'),
        #         right=Side(border_style=BORDER_THIN, color='00000000'),
        #         top=Side(border_style=BORDER_THIN, color='00000000'),
        #         bottom=Side(border_style=BORDER_THIN, color='00000000')
        #     )
        #     alignment = Alignment(horizontal='center',
        #                           vertical='center',
        #                           text_rotation=0,
        #                           wrap_text=True,
        #                           shrink_to_fit=True,
        #                           indent=0)
        #     data_objects = json.loads(response.text)
        #     sheet.cell(row=2, column=2).value = querystring
        #     for data in data_objects['objects']:
        #         print("data ai: ",data)
        #         sheet.cell(row=cur_row, column=1).value = str(index)
        #         sheet.cell(row=cur_row, column=2).value = data['patent_code']
        #         sheet.cell(row=cur_row, column=3).value = data['patent_paragraph']
        #         sheet.cell(row=cur_row, column=1).border = thin_border
        #         sheet.cell(row=cur_row, column=2).border = thin_border
        #         sheet.cell(row=cur_row, column=3).border = thin_border
        #         sheet.cell(row=cur_row, column=1).alignment = alignment
        #         sheet.cell(row=cur_row, column=2).alignment = alignment
        #         sheet.cell(row=cur_row, column=3).alignment = alignment
        #         cur_row += 1
        #         index += 1
        #     wb.save(filename=new_xls_file)
        return make_response(response.text, 200)
    except :
        pass
    return make_response(jsonify(''), 200)
#
# @app.route('/api/v1/check_kitech_patent', methods=['GET'])
# def check_kitech_patent_api():
#     print("check_kitech_patent start")
#     start_date = request.args.get('start_date')
#     end_date = request.args.get('end_date')
#
#     def add_kitech_source(patent_code,invention_title,description,description_of_embodiments):
#         new_kitech_source = KitechSource()
#         new_kitech_source.patent_code = patent_code
#         new_kitech_source.invention_title = invention_title
#         new_kitech_source.description = description
#         new_kitech_source.check_ = 1
#         new_kitech_source.description_of_embodiments = description_of_embodiments
#         db.session.add(new_kitech_source)
#         db.session.commit()
#
#     def insert_patent_source(patent_code,p_idx,k_nouns_p,invention_title_noun,patent_paragraph,patent_paragraph_expend) :
#         new_patent_paragraph1 = TB_KITECH_PATENT_PARAGRAPH()
#         new_patent_paragraph1.patent_code = patent_code
#         new_patent_paragraph1.p_idx = p_idx
#         new_patent_paragraph1.k_nouns_p = k_nouns_p
#         new_patent_paragraph1.invention_title_noun = invention_title_noun
#         new_patent_paragraph1.patent_paragraph = patent_paragraph
#         new_patent_paragraph1.patent_paragraph_expend = patent_paragraph_expend
#         db.session.add(new_patent_paragraph1)
#         db.session.commit()
#
#     def check_text(strSource,nLen):
#         doc_len = len(strSource)
#         if doc_len > nLen:
#             return  strSource[:nLen]
#         return strSource
#
#     ha = Hannanum()
#     headers = {'User-Agent': 'Mozilla/5.0'}
#     date_duration = start_date + "~" + end_date
#     payload = 'remoconExpression=&remoconDocsFound=&remoconSelectedArticles=&queryText=GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D&searchInResultCk=undefined&next=MainList&passwd=&userId=&config=G1111111111111111111111S111111111000000000&sortField=RANK&sortState=DESC&sortField1=&sortState1=&sortField2=&sortState2=&configChange=Y&expression=((GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D)*(P%3CIN%3ECLS))*(%7BR%3CIN%3ELST%7D)&historyQuery=GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D&numPerPage=90&numPageLinks=10&currentPage=1&FROM=&BOOKMARK=&NWBOOKMARK=&REBOOKMARK=&natlCD=&beforeExpression=((GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D)*(P%3CIN%3ECLS))*(%7BR%3CIN%3ELST%7D)&userInput=%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0&searchKeyword=&searchInTrans=N&searchInResult=&logFlag=Y&searchSaveCnt=0&piField=&piValue=&piSearchYN=N&leftGubnChk=&leftHangjungChk=&SEL_PAT=KPAT&merchandiseString=&measureString=&patternString=&collectionValues=&selectedLang=&lang=&strstat=SMART%7CGD%7CNDN%7C&searchInTransCk=undefined'.format(date_duration)
#     print("payload : ",payload)
#     session = requests.Session()
#     r = session.post('http://kpat.kipris.or.kr/kpat/resulta.do',data=payload,headers={'Content-Type': 'application/x-www-form-urlencoded'})
#
#     # print("response : ",r.text)
#     soup = BeautifulSoup(r.text, 'html.parser')
#     doc_len = 0
#     patent_list = soup.find_all('article')
#     print("count : ",len(patent_list))
#     new_patent_list = []
#     dest_thumb = app.config["STATIC_FILE_PATH"] + "thumbnail/"
#     dest_pdf = app.config["STATIC_FILE_PATH"] + "pdf/"
#     for article in patent_list:
#         patent_code = article.get('id').replace('divViewSel','')
#         print("patent code :", patent_code)
#         source = KitechSource.query.filter_by(patent_code=patent_code).first()
#         if source is None :
#             print("add patent code :", patent_code)
#             invention_title = ''
#             summary_of_invention = ''
#             new_patent_list.append(patent_code)
#             print("must be new patent code :", patent_code)
#             invention_title = article.find('input').get('title')
#             # print("invention_title :",invention_title)
#             for div in article.find_all('div') :
#                 class_name = div.get('class')[0]
#                 # print("div class : ",class_name)
#                 if class_name == 'search_txt':
#                     # print("div contents : ", div.contents[2])
#                     summary_of_invention = re.sub('<.+?>', '', str(div.contents[2]), 0).strip()
#                     break
#             # print("summary_of_invention :", summary_of_invention)
#             description_list = []
#             code = ''
#             if patent_code[:2] == '10' :
#                 description_list.append('"KR' + patent_code[2:] + 'A"')
#             else :
#                 description_list.append('"KR' + patent_code[2:] + 'U"')
#             description_list.append('"' + invention_title + '"')
#             description_list.append('"' + summary_of_invention + '"')
#             description = ','.join(description_list)
#             # print("description :",description)
#
#             invention_title_noun_ha = ha.nouns(invention_title[:500].replace('(','').replace(')',''))
#
#             url = "http://kpat.kipris.or.kr/kpat/remoteFile.do?method=downloadImage&applno={0}".format(patent_code)
#             response = requests.get(url, stream=True)
#             file_name = dest_thumb + patent_code + '.jpg'
#             print("file_name : ", file_name)
#             with open(file_name, "wb") as handle:
#                 for data in tqdm(response.iter_content()):
#                     handle.write(data)
#             url = "http://kpat.kipris.or.kr/kpat/{0}.pdf?method=fullText&applno={0}".format(patent_code)
#             response = requests.get(url, stream=True)
#             file_name = dest_pdf + patent_code + '.pdf'
#             print("file_name : ", file_name)
#             with open(file_name, "wb") as handle:
#                 for data in tqdm(response.iter_content()):
#                     handle.write(data)
#             parsed = parser.from_file(file_name)
#             find_index = parsed["content"].find('발명의 설명')
#             if find_index > 0 :
#                 docs =re.sub('<.+?>', '', parsed["content"][find_index:], 0).strip().replace('\n',' ')
#                 docs_list  = []
#                 short_doc = ''
#                 print("doc LEN1 : ", len(docs))
#                 # docs = self.check_text(docs, 1024 * 1000 * 2)
#                 # print("doc LEN2 : ", len(docs))
#                 # print("docs : ", docs[:200])
#                 docs = docs[:20000]
#                 add_kitech_source(patent_code,invention_title,description,docs[:30000])
#
#                 for doc in docs.split('. ') :
#                     if len(doc) < 200 :
#                         short_doc += ' ' + doc
#                     elif short_doc != '' and len(doc) >= 200:
#                         doc = short_doc + ' ' + doc
#                         docs_list.append(doc)
#                         short_doc = ''
#                         # print("doc : ", doc)
#                     else :
#                         # print("doc : ", doc)
#                         docs_list.append(doc)
#                         short_doc = ''
#
#                 print("docs :",len(docs_list))
#                 cur = 1
#                 for m_desc in docs_list :
#                     paragraph_noun_ha = ha.nouns(m_desc[:1024])
#                     # print("paragraph_noun_ha start 1")
#                     paragraph_noun_ha = paragraph_noun_ha[:510]
#                     if len(paragraph_noun_ha) > 0:
#                         # print("paragraph_noun_ha 0")
#                         df2 = pd.Series(paragraph_noun_ha)
#                         # print("paragraph_noun_ha 1")
#                         if len(invention_title_noun_ha) > 0:
#                             k_noun_count_list = [df2.str.count(noun).sum() for noun in invention_title_noun_ha]
#                             # print("paragraph_noun_ha 2")
#                             np.sum(k_noun_count_list)
#                             # print("paragraph_noun_ha 3")
#                             # print("=======> len :", len(m_desc), ", desc :", m_desc)
#                             desc_expend = invention_title + " " + m_desc
#                             insert_patent_source(patent_code, cur, np.sum(k_noun_count_list),invention_title, m_desc,desc_expend)
#                             # print("paragraph_noun_ha 4")
#                     cur += 1
#         print("skip pagent : ", patent_code)
#     print("must be count : ", len(new_patent_list))
#     return make_response(jsonify(''), 200)


@app.route('/api/v1/check_kitech_patent2', methods=['GET'])
def check_kitech_patent_api():
    print("check_kitech_patent start")
    start_date = "20200101"
    end_date = "20201231"

    ha = Hannanum()
    headers = {'User-Agent': 'Mozilla/5.0'}
    date_duration = start_date + "~" + end_date
    payload = 'remoconExpression=&remoconDocsFound=&remoconSelectedArticles=&queryText=GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D&searchInResultCk=undefined&next=MainList&passwd=&userId=&config=G1111111111111111111111S111111111000000000&sortField=RANK&sortState=DESC&sortField1=&sortState1=&sortField2=&sortState2=&configChange=Y&expression=((GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D)*(P%3CIN%3ECLS))*(%7BR%3CIN%3ELST%7D)&historyQuery=GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D&numPerPage=90&numPageLinks=10&currentPage=1&FROM=&BOOKMARK=&NWBOOKMARK=&REBOOKMARK=&natlCD=&beforeExpression=((GD%3D%5B{0}%5D*NDN%3D%5B%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0%5D)*(P%3CIN%3ECLS))*(%7BR%3CIN%3ELST%7D)&userInput=%EC%83%9D%EC%82%B0%EA%B8%B0%EC%88%A0&searchKeyword=&searchInTrans=N&searchInResult=&logFlag=Y&searchSaveCnt=0&piField=&piValue=&piSearchYN=N&leftGubnChk=&leftHangjungChk=&SEL_PAT=KPAT&merchandiseString=&measureString=&patternString=&collectionValues=&selectedLang=&lang=&strstat=SMART%7CGD%7CNDN%7C&searchInTransCk=undefined'.format(date_duration)
    print("payload : ",payload)
    session = requests.Session()
    r = session.post('http://kpat.kipris.or.kr/kpat/resulta.do',data=payload,headers={'Content-Type': 'application/x-www-form-urlencoded'})

    # print("response : ",r.text)
    soup = BeautifulSoup(r.text, 'html.parser')
    doc_len = 0
    patent_list = soup.find_all('article')
    print("count : ",len(patent_list))
    new_patent_list = []
    dest_thumb = app.config["STATIC_FILE_PATH"] + "thumbnail/"
    dest_pdf = app.config["STATIC_FILE_PATH"] + "pdf/"
    for article in patent_list:
        patent_code = article.get('id').replace('divViewSel','')
        print("article :", article)
        print("patent_code :", patent_code)

    return make_response(jsonify(''), 200)

#
#
# @app.route('/api/v1/check_kicet_patent', methods=['GET'])
# def check_kicet_patent_api():
#     print("check_kicet_patent start")
#
#     def insert_patent_source(patent_code,p_idx,k_nouns_p,invention_title_noun,patent_paragraph,patent_paragraph_expend) :
#         new_patent_paragraph1 = TB_KICET_PATENT_PARAGRAPH()
#         new_patent_paragraph1.patent_code = patent_code
#         new_patent_paragraph1.p_idx = p_idx
#         new_patent_paragraph1.k_nouns_p = k_nouns_p
#         new_patent_paragraph1.invention_title_noun = invention_title_noun
#         new_patent_paragraph1.patent_paragraph = patent_paragraph
#         new_patent_paragraph1.patent_paragraph_expend = patent_paragraph_expend
#         db.session.add(new_patent_paragraph1)
#         db.session.commit()
#
#     def check_text(strSource,nLen):
#         doc_len = len(strSource)
#         if doc_len > nLen:
#             return  strSource[:nLen]
#         return strSource
#
#     ha = Hannanum()
#     headers = {'User-Agent': 'Mozilla/5.0'}
#     new_patent_list = []
#     dest_thumb = "D:\\project\\patent\\download\\thumbnail\\"
#     dest_pdf = "D:\\project\\patent\\download\\thumbnail\\pdf\\"
#     temp_list = TB_TEMP.query.all()
#     for temp in temp_list:
#         patent_code = temp.patent_code
#         print("patent code :", patent_code)
#         source = TB_KICET_SOURCE.query.filter_by(patent_code=patent_code).first()
#         if source is not None :
#             print("add patent code :", patent_code)
#             invention_title = source.invention_title
#             invention_title_noun_ha = ha.nouns(invention_title[:500].replace('(','').replace(')',''))
#             url = "http://kpat.kipris.or.kr/kpat/remoteFile.do?method=downloadImage&applno={0}".format(patent_code)
#             response = requests.get(url, stream=True)
#             file_name = dest_thumb + patent_code + '.jpg'
#             print("file_name : ", file_name)
#             with open(file_name, "wb") as handle:
#                 for data in tqdm(response.iter_content()):
#                     handle.write(data)
#             url = "http://kpat.kipris.or.kr/kpat/{0}.pdf?method=fullText&applno={0}".format(patent_code)
#             response = requests.get(url, stream=True)
#             file_name = dest_pdf + patent_code + '.pdf'
#             print("file_name : ", file_name)
#             with open(file_name, "wb") as handle:
#                 for data in tqdm(response.iter_content()):
#                     handle.write(data)
#             print("parsed start")
#             parsed = parser.from_file(file_name)
#             print("parsed ",parsed)
#             find_index = parsed["content"].find('발명의 설명')
#             print("parsed find_index 1: ", find_index)
#             if find_index == -1 :
#                 find_index = parsed["content"].find('발명의 상세한 설명')
#             print("parsed find_index 2: ", find_index)
#             docs = ''
#             docs = re.sub('<.+?>', '', parsed["content"], 0).strip().replace('\n', ' ')
#             docs_list  = []
#             short_doc = ''
#             # print("doc LEN1 : ", len(docs))
#             # docs = self.check_text(docs, 1024 * 1000 * 2)
#             # print("doc LEN2 : ", len(docs))
#             print("docs : ", docs)
#             docs = docs[:20000]
#             for doc in docs.split('. ') :
#                 if len(doc) < 200 :
#                     short_doc += ' ' + doc
#                 elif short_doc != '' and len(doc) >= 200:
#                     doc = short_doc + ' ' + doc
#                     docs_list.append(doc)
#                     short_doc = ''
#                     # print("doc : ", doc)
#                 else :
#                     # print("doc : ", doc)
#                     docs_list.append(doc)
#                     short_doc = ''
#
#             print("docs :",len(docs_list))
#             cur = 1
#             for m_desc in docs_list :
#                 paragraph_noun_ha = ha.nouns(m_desc[:1024])
#                 # print("paragraph_noun_ha start 1")
#                 paragraph_noun_ha = paragraph_noun_ha[:510]
#                 if len(paragraph_noun_ha) > 0:
#                     # print("paragraph_noun_ha 0")
#                     df2 = pd.Series(paragraph_noun_ha)
#                     # print("paragraph_noun_ha 1")
#                     if len(invention_title_noun_ha) > 0:
#                         k_noun_count_list = [df2.str.count(noun).sum() for noun in invention_title_noun_ha]
#                         # print("paragraph_noun_ha 2")
#                         np.sum(k_noun_count_list)
#                         # print("paragraph_noun_ha 3")
#                         # print("=======> len :", len(m_desc), ", desc :", m_desc)
#                         desc_expend = invention_title + " " + m_desc
#                         insert_patent_source(patent_code, cur, np.sum(k_noun_count_list),invention_title, m_desc,desc_expend)
#                         # print("paragraph_noun_ha 4")
#
#                 cur += 1
#         print("skip pagent : ", patent_code)
#     print("must be count : ", len(new_patent_list))
#     return make_response(jsonify(''), 200)

manager.create_api(KitechReport
                   , url_prefix='/api/v1'
                   , collection_name='report_search_hist'
                   , methods=['GET', 'PUT', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [check_token],
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token]
                   })

def post_get_question(result=None, **kw):
    check_token()

    src_list = result['objects']

    for src in src_list:
        user = Users.query.filter_by(user_id=src['user_id']).first()
        if user is not None :
            src['email'] = user.email
            src['phone'] = user.phone

manager.create_api(KITECH_QUESTION
                   , url_prefix='/api/v1'
                   , collection_name='question'
                   , methods=['GET', 'PUT', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [check_token],
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token]
                   },
                   postprocessors={
                       'GET_MANY': [post_get_question]
                   })

@app.route('/api/v1/searchhist', methods=['DELETE'])
def searchhist_api():
    print("searchhist_api")
    vendor = TB_VENDOR.query.filter_by(active=1).first()
    db.session.query(TB_KITECH_SEARCH_HIST).filter_by(fk_vendor_id=vendor.id).delete()
    db.session.commit()
    return make_response(jsonify(''), 200)

@app.route('/api/v1/sensor_data_upload', methods=['POST'])
def sensor_data_upload():
    print("sensor_data_upload")
    result = {
        "ACK":"SUCCESS",
        "MSG":""
    }
    try :
        parser = reqparse.RequestParser()
        parser.add_argument("sender_ip", type=str, location="json")
        parser.add_argument("company_id", type=str, location="json")
        parser.add_argument("sensors_msg", type=str, location="json")
        input = parser.parse_args()
        print("input : ", input)
        #client = MongoClient('localhost', 27017)
        #db = client['itrc_spark']
        #collection = db.col_sensor
        obj = {
            "sender_ip": input['sender_ip'],
            "company_id": input['company_id'],
            "sensors_msg": json.dumps(input['sensors_msg']),
        }

        new_msg = TB_SENSOR_UPLOAD()
        new_msg.sender_ip = input['sender_ip']
        new_msg.company_id = input['company_id']
        new_msg.sensors_msg = json.dumps(input['sensors_msg'])
        new_msg.created_date = datetime.now()
        db.session.add(new_msg)
        db.session.commit()
    #collection.insert(obj)
    except :
        result = {
            "ACK":"FAIL",
            "MSG":"mongodb input exception"
        }

    return make_response(jsonify(result), 200)

none_list = []
except_list = []
# import pdfplumber
@app.route('/api/v1/parse_and_make_txt', methods=["GET"])
def parse_and_make_txt_api():
    def make_txt(file_path,names):
        global none_list
        global except_list

        file_name = file_path + "\\" + names
        try:
            pdf = pdfplumber.open(file_name)
        except:
            except_list.append(names)
            return
        text = pdf.pages[0].extract_text()
        if text == None:
            none_list.append(names)
            return None

        with open('C:\\d\\project\\kitech_prj\\txts\\'+names.replace('.pdf','.txt'),'wb') as txt_file:
            txt_file.write(text.encode())

        return None

    print('pdf_parse_api ==> run')
    file_path = 'C:\\d\\project\\kitech_prj\\pdf'
#     file_name = 'C:\\d\\project\\kitech_prj\\pdf\\1019840001881.pdf'

    file_list = os.listdir(file_path)
    for cur,names in enumerate(file_list):
        print('cur :',cur,'path :',names)
        make_txt(file_path,names)

    with open('C:\\d\\project\\kitech_prj\\'+'nones.txt','wb') as nones:
        for nl in none_list:
            nones.write(nl)
    with open('C:\\d\\project\\kitech_prj\\'+'excepts.txt','wb') as excepts:
        for exs in except_list:
            excepts.write(exs)


    print('api end')

    return make_response(jsonify(), 200)


# import shutil
@app.route('/api/v1/txt_count', methods=["GET"])
#못읽은 pdf들 다른 폴더로 이동시키기
def txt_count_api():
    print('txt_count_api ==> run')

    file_path = 'C:\\d\\project\\kitech_prj'
    #1. another
#     pdfs = os.listdir(file_path+'\\pdf')
#     txts = os.listdir(file_path+'\\txts')
#     txts = [t.replace('.txt','.pdf') for t in txts]
#     print('pdf len :',len(pdfs),' /  txt len :',len(txts))
#     print(pdfs[0],txts[0])
#     true_list = []
#
#     for cur,txt in enumerate(txts):
#         if txt in pdfs:
#             true_list.append(txt)
#         if cur%1000 == 0:
#             print('cur :',cur)
#
#     for trues in true_list:
#         pdfs.remove(trues)
#         txts.remove(trues)
#
#     for pdf in pdfs:
#         shutil.move(file_path+'\\pdf\\'+pdf,file_path+'\\pdf_another\\'+pdf)
#
#     print('len :',len(pdfs))

    # empty
    empty_prs_txt = open(file_path+'\\pdf_test\\empty_prs.txt','r')
    empty_list = empty_prs_txt.readlines()
    for cur,em in enumerate(empty_list):
        file_name = em.replace('\n','.pdf')
        shutil.move(file_path+'\\pdf\\'+file_name,file_path+'\\pdf_empty\\'+file_name)
        print(cur)

    print('api end')
    return make_response(jsonify(), 200)


# error_list = []
@app.route('/api/v1/pdf_parse', methods=["GET"])
def pdf_parse_api():
    def pdf_parse(file_path,names,data_dict):
#         global error_list
#         global data_dict
        file_name = file_path + "\\" + names
        pdf = pdfplumber.open(file_name)
        text = pdf.pages[0].extract_text()
        if text == None:
#             error_list.append(names)
            return print('None')

        find_word = text.find('출원인')
        cat = 0
        if find_word == -1:
            find_word = text.find('특허권자')
            cat = 1
            if find_word == -1:
                find_word = text.find('실용신안권자')
                cat = 2
#                 if find_word == -1:
#                     error_list.append(file_name)
        find_ent = text.find('\n',find_word+1)
        find_txt = text[find_word:find_ent]
#         if find_txt == '':
#             data_dict[names.replace('.pdf','')] = find_txt
        if cat == 1:
            txt = find_txt.replace('특허권자','')
            txt_list = txt.split(' ')
            txt_list = [t for t in txt_list if t!=' ' and t!='']
            txt = ' '.join(txt_list)
            data_dict[names.replace('.pdf','')] = txt
        elif cat == 2:
            txt = find_txt.replace('실용신안권자','')
            txt_list = txt.split(' ')
            txt_list = [t for t in txt_list if t!=' ' and t!='']
            txt = ' '.join(txt_list)
            data_dict[names.replace('.pdf','')] = txt
        else:
            txt = find_txt.replace('출원인','')
            txt_list = txt.split(' ')
            txt_list = [t for t in txt_list if t!=' ' and t!='']
            txt = ' '.join(txt_list)
            data_dict[names.replace('.pdf','')] = txt

    print('pdf_parse_api ==> run')
    file_path = 'C:\\d\\project\\kitech_prj\\pdf'
#     error_list = []
    data_dict = {}
#     file_name = 'C:\\d\\project\\kitech_prj\\pdf\\1019840001881.pdf'

    file_list = os.listdir(file_path)
    for cur,names in enumerate(file_list):
        print('cur :',cur,'path :',names)
        pdf_parse(file_path,names,data_dict)

    print('parse end :',len(data_dict))
#     print(error_list)

    data = KitechSource.query.all()
    cur = 0
    error_list = []
    for ids,prs in data_dict.items():
#         obj = {'person':prs}

#         db.session.query(KitechSource).filter_by(patent_code=ids).update(obj)
        cur+=1
        if prs == '':
            print('cur :',cur)
            error_list.append(ids)
#     db.session.commit()

    with open('C:\\d\\project\\kitech_prj\\pdf_test\\empty_prs.txt','wb') as empty_prs:
        for errors in error_list:
            er_txt = errors+'\n'
            empty_prs.write(er_txt.encode())

    print('dict len :', len(data_dict),'cur :',cur)
    print('api end')

    return make_response(jsonify(), 200)


@app.route('/api/v1/pdf_parse2', methods=["GET"])
def pdf_parse2_api():
    print('pdf_parse2_api ==> run')
    data_dict = {}
    empty_list = []

    file_path = 'C:\\d\\project\\kitech_prj'
    path_list = ['\\pdf_another','\\pdf_empty','\\pdf_empty2']
    path = file_path+path_list[2]
#     path2 = file_path+path_list[2]
    pdfs = os.listdir(path)
    for cur,pdf in enumerate(pdfs):

        file_name = path + '\\' + pdf
        print('cur :',cur)
        text = parser.from_file(file_name)
        text = text['content']

        find_word = text.find('출원인')
        cat = 0
        if find_word == -1:
            find_word = text.find('특허권자')
            cat = 1
            if find_word == -1:
                find_word = text.find('실용신안권자')
                cat = 2

        find_ent = text.find('\n',find_word+1)
        find_txt = text[find_word:find_ent]
        find_txt = ' '.join(find_txt.split(' ')[1:])
        print(find_txt)
        break

        if find_txt == '':
            empty_list.append(pdf.replace('.pdf',''))
            continue
        data_dict[pdf.replace('.pdf','')] = find_txt


    print(len(data_dict))
    print(len(empty_list))

#     data = KitechSource.query.all()
#     for ids,prs in data_dict.items():
#         obj = {'person':prs}
#         db.session.query(KitechSource).filter_by(patent_code=ids).update(obj)
#     db.session.commit()
#     print('update')

#     for cur,em in enumerate(empty_list):
#         file_name = path+em+'.pdf'
#         file_name2 = path2+em+'.pdf'
#         shutil.move(file_name,file_name2)
#     for ids,prs in data_dict.items():
#         file_name = path+'\\'+ids+'.pdf'
#         file_name2 = path2+'\\'+ids+'.pdf'
#         print(file_name)
#         print(file_name2)
#         shutil.move(file_name,file_name2)
    print('move')


#     print(data_dict)
#     print(empty_list)

    print('api end')
    return make_response(jsonify(), 200)

@app.route('/api/v1/pdf_parse2_plum', methods=["GET"])
def pdf_parse2_plum_api():
    print('pdf_parse2_plum_api ==> run')
    data_dict = {}
    empty_list = []

    file_path = 'C:\\d\\project\\kitech_prj'
    path_list = ['\\pdf_another','\\pdf_empty','\\pdf_empty2']
    path = file_path+path_list[2]
#     path2 = file_path+path_list[2]
    pdfs = os.listdir(path)
    for cur,pdf in enumerate(pdfs):

        file_name = path + '\\' + pdf
        pdf = pdfplumber.open(file_name)
        text = pdf.pages[0].extract_text()
        print(file_name)
        print('cur :',cur)

        find_word = text.find('출원인')
        cat = 0
        if find_word == -1:
            find_word = text.find('특허권자')
            cat = 1
            if find_word == -1:
                find_word = text.find('실용신안권자')
                cat = 2

        find_ent = text.find('\n',find_word+1)
        find_txt = text[find_word:find_ent]
        find_txt = ' '.join(find_txt.split(' ')[1:])
        print(find_txt)
        break

        if find_txt == '':
            empty_list.append(pdf.replace('.pdf',''))
            continue
        data_dict[pdf.replace('.pdf','')] = find_txt


#     print(len(data_dict))
#     print(len(empty_list))

#     data = KitechSource.query.all()
#     for ids,prs in data_dict.items():
#         obj = {'person':prs}
#         db.session.query(KitechSource).filter_by(patent_code=ids).update(obj)
#     db.session.commit()
#     print('update')

#     for cur,em in enumerate(empty_list):
#         file_name = path+em+'.pdf'
#         file_name2 = path2+em+'.pdf'
#         shutil.move(file_name,file_name2)
#     for ids,prs in data_dict.items():
#         file_name = path+'\\'+ids+'.pdf'
#         file_name2 = path2+'\\'+ids+'.pdf'
#         print(file_name)
#         print(file_name2)
#         shutil.move(file_name,file_name2)
    print('move')


#     print(data_dict)
#     print(empty_list)

    print('api end')
    return make_response(jsonify(), 200)

@app.route('/api/v1/web_txt_parse', methods=["GET"])
def web_txt_parse_api():
    print('pdf_parse2_plum_api ==> run')
    def roof1(num):
        print('roof1 run :',num)

        file_path = 'C:\\d\\project\\kitech_prj\\pages\\{0}페이지.txt'.format(num)
        page7 = open(file_path,'rb')
        soup = BeautifulSoup(page7, 'html.parser')
        sp = soup.find_all('article')
        print(len(sp))
        for cur,sps in enumerate(sp):
            sp1 = sps.find('li', attrs={'class':'right_width'})
            sp1_txt = sp1.find('font').text
            sp2 = sps.find_all('li', attrs={'class':'left_width'})
            sp2_ = [tags.find('a').text if tags.find('a')!=None else None for tags in sp2]
            sp2__ = sp2_[1]
            sp2_txt = sp2__.split(' ')[0]
            print(cur,'/ person :',sp1_txt,'/ patent_code :',sp2_txt)
            obj = {'person':sp1_txt}
            db.session.query(KitechSource).filter_by(patent_code=sp2_txt).update(obj)
            if cur%100==0:
                db.session.commit()
        db.session.commit()

    for num in range(7):
        roof1(str(num+1))

    print('api end')
    return make_response(jsonify(), 200)

@app.route('/api/v1/pdf_down', methods=["GET"])
def pdf_down_api():
    print('pdf_down_api ==> run')

    dest_thumb = "C:\\d\\project\\kitech_prj\\down_img\\"
    dest_pdf = "C:\\d\\project\\kitech_prj\\down_pdf\\"
    downs = db.session.query(KitechSource).all()
    for cur,db_data in enumerate(downs):
        if cur<432:
            continue
        patent_code = db_data.patent_code

        url = "http://kpat.kipris.or.kr/kpat/remoteFile.do?method=downloadImage&applno={0}".format(patent_code)
        response = requests.get(url, stream=True)
        file_name = dest_thumb + patent_code + '.jpg'
        print("file_name : ", file_name)
        with open(file_name, "wb") as handle:
            for data in tqdm(response.iter_content()):
                handle.write(data)
        url = "http://kpat.kipris.or.kr/kpat/{0}.pdf?method=fullText&applno={0}".format(patent_code)
        response = requests.get(url, stream=True)
        file_name = dest_pdf + patent_code + '.pdf'
        print("file_name : ", file_name)
        with open(file_name, "wb") as handle:
            for data in tqdm(response.iter_content()):
                handle.write(data)
        print('\ncur :',cur)

    print('api end')
    return make_response(jsonify(), 200)

@app.route('/api/v1/source_no_exis', methods=["GET"])
def source_no_exis_api():
    print('source_no_exis_api ==> run')
    def roof1(num):
        print('roof1 run :',num)

        file_path = 'C:\\d\\project\\kitech_prj\\pages\\{0}페이지.txt'.format(num)
        page7 = open(file_path,'rb')
        soup = BeautifulSoup(page7, 'html.parser')
        sp = soup.find_all('article')
        for cur,sps in enumerate(sp):
            sp1 = sps.find_all('li', attrs={'class':'left_width'})
            sp1_ = [tags.find('a').text if tags.find('a')!=None else None for tags in sp1]
            src_code = sp1_[1].split(' ')[0]
#             db_exis = db.session.query(KitechSource).filter_by(patent_code=src_code).first()
#             if db_exis!=None:
#                 continue
            sp2 = sps.find(attrs={'class':'search_section_title'})
            src_title = sp2.find_all('a')[1].text.replace('\t','').replace('\r','').replace('\n',' ')
            sp3 = sps.find('li', attrs={'class':'right_width'})
            src_person = sp3.find('font').text
            sp4 = sps.find(attrs={'class':'search_txt'})
            code_cat = 'U'
            if src_code[0] == '1':
                code_cat = 'A'
            txt1 = "KR{0}{1}".format(src_code[2:],code_cat)
            txt2 = sp4.text.strip().replace('\t','').replace('\r','').replace('\n',' ')
            src_text = '\"{0}\",\"{1}\",\"{2}\"'.format(txt1,src_title,txt2)
            print('\n',cur)
            print('patent_code :',src_code,'/ title :',src_title)
            print('person :',src_person,'/ content :',src_text)

            src = KitechSource()
            src.patent_code = src_code
            src.invention_title = src_title
            src.person = src_person
            src.description = src_text
            db.session.add(src)
            db.session.commit()

    for num in range(7):
        roof1(str(num+1))

    print('api end')
    return make_response(jsonify(), 200)

import pdfplumber
error_list = []
@app.route('/api/v1/pdf_paragraph', methods=["GET"])
def pdf_paragraph_api():
    def roof1(file_name):
        global error_list
        pdf_path = 'C:\\d\\project\\kitech_prj\\down_pdf'
        file_path = pdf_path+'\\{0}.pdf'.format(file_name)

        pdf = pdfplumber.open(file_path)
        docs1 = [doc.extract_text().replace('\n','') for doc in pdf.pages if doc.extract_text()!=None]
        docs = ''.join(docs1)

        db_data = db.session.query(KitechSource).filter(KitechSource.patent_code.like("%"+file_name+"%")).first()
        if db_data==None:
            db_title = ''
        else:
            db_title = db_data.invention_title

        find_index = docs.find('발명의 설명')
        txt_cat = 0
        if find_index == -1 :
            find_index = docs.find('발명의 상세한 설명')
            txt_cat = 1
            if find_index == -1:
                find_index = docs.find('요약')
                txt_cat = 2
                if find_index == -1:
                    error_list.append(file_name)
                    return

        if txt_cat==0:
            text = docs[find_index+7:find_index+20007]
        elif txt_cat==0:
            text = docs[find_index+11:find_index+20011]
        else:
            text = docs[find_index+2:find_index+20002]
        para_list = text.split('. ')
        docs_list = []
        short_doc=''
        for doc in para_list :
            if len(doc) < 200 :
                short_doc += ' ' + doc
            elif short_doc != '' and len(doc) >= 200:
                doc = short_doc + '. ' + doc
                docs_list.append(doc)
                short_doc = ''
            else :
                docs_list.append(doc)
                short_doc = ''

        if len(docs_list) > 0:
            for cur,txt in enumerate(docs_list):
                para = TB_KITECH_PATENT_PARAGRAPH()
                para.patent_code = file_name
                para.p_idx = cur+1
                para.k_nouns_p = 0
                para.invention_title_noun = db_title
                para.patent_paragraph = txt
                para.patent_paragraph_expend = db_title+' '+txt
                db.session.add(para)
                db.session.commit()


    print('pdf_paragraph_api run')
    global error_list
    error_list = []

    file_list = os.listdir('C:\\d\\project\\kitech_prj\\down_pdf')
    for cur,files in enumerate(file_list):
        file_name = files.replace('.pdf','')
        roof_filter = db.session.query(TB_KITECH_PATENT_PARAGRAPH).filter(TB_KITECH_PATENT_PARAGRAPH.patent_code.like("%"+file_name+"%")).first()
        if roof_filter==None:
            roof1(file_name)
        print('cur :',cur)

    for ers in error_list:
        shutil.move('C:\\d\\project\\kitech_prj\\down_pdf\\{0}.pdf'.format(ers),'C:\\d\\project\\kitech_prj\\test\\{0}.pdf'.format(ers))

    print('api end')
    return make_response(jsonify({}), 200)

import shutil
@app.route('/api/v1/no_inserts', methods=["GET"])
def no_inserts_api():
    print('pdf_paragraph_api run')
    query_txt = "SELECT * FROM tb_kitech_source WHERE patent_code NOT IN (SELECT patent_code FROM tb_kitech_patent_paragraph)"
    db_datas = db.session.execute(sql.text(query_txt))

    file_path = 'C:\\d\\project\\kitech_prj'
    for db_data in db_datas:
        file_name = db_data.patent_code
        old = file_path+'\\down_pdf\\'+file_name+'.pdf'
        new = file_path+'\\test\\'+file_name+'.pdf'
        shutil.move(old,new)

    print('api end')
    return make_response(jsonify({}), 200)

manager.create_api(Community
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='community'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [check_token],
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token_single],
                        'GET_MANY': [check_token]
                   })

manager.create_api(SmallPatent
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='smallpatent'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                   })

# manager.create_api(KitechSource
#                    , results_per_page=10000
#                    , url_prefix='/api/v1'
#                    , collection_name='sources'
#                    , methods=['GET', 'DELETE', 'PATCH', 'POST']
#                    , allow_patch_many=True
#                    , preprocessors={
#                         'POST': [check_token],
#                         'PATCH_SINGLE': [check_token],
#                         'GET_SINGLE': [check_token],
#                    })

def post_kitech_source(result=None, **kw):
    print("patent_code :",result['patent_code'])
    if result['patent_code'] is not None :
        patent_pdf_file_path = os.getcwd() + '/pdf/' + result['patent_code'] + ".pdf"
        if os.path.exists(patent_pdf_file_path):
            result['patent_pdf_file_path'] = result['patent_code'] + ".pdf"
    if result['maketting_file_path'] is not None :
        patent_maketting_file_path = os.getcwd() + '/uploads/' + result['maketting_file_path']
        if os.path.exists(patent_maketting_file_path):
            result['patent_maketting_file_path'] = result['maketting_file_path']
    if result['patent_code'] is not None :
        patent_thumb_image_path = os.getcwd() + '/thumbnail/' + result['patent_code'] + ".jpg"
        if os.path.exists(patent_thumb_image_path):
            result['patent_thumb_image_path'] = result['patent_code'] + ".jpg"

manager.create_api(KitechSource
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='sources'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , postprocessors={'GET_SINGLE': [post_kitech_source]}
                   , preprocessors={
                        'POST': [check_token],
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token_single],
                   })
manager.create_api(PopupZone
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='popupzone'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                        'POST': [check_token],
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token_single],
                   })

def post_tech_trans(result=None, **kw):
    print("result : ", result)
    message = '''<pre>
        기술이전 요청이 접수되었습니다.
        ************* 기술이전 정보 ****************
        기업명 : {0}
        특허출원번호 : {1}
        요청자이름 : {2}
        연락처 : {3}
        이메일 : {4}
        요청사항 : {5}
        ******************************************
            </pre>'''.format(
                result['company_name'],
                result['patent_code'],
                result['request_name'],
                result['request_phone'],
                result['request_email'],
                result['request_contents'],
            )
    print("message :",message)
    server = smtplib.SMTP_SSL(app.config['SMTP_ADDR'], app.config['SMTP_PORT'])
    server.login(app.config['SMTP_LOGIN_ID'], app.config['SMTP_LOGIN_PW'])
    msg = MIMEMultipart('alternative')
    msg['From'] = "%s <%s>" % ("", app.config['SMTP_SENDER'])
    msg['To'] = 'honsu@kitech.re.kr'
#     msg['To'] = 'akj2995@daum.net'
    msg['Subject'] = "[생산기술연구원-인공지능특허검색시스템] 기술이전 정보 접수1"
    msg.attach(MIMEText(message, 'html', 'utf-8'))  # 내용 인코딩
    server.sendmail(app.config['SMTP_SENDER'], msg['To'], msg.as_string())

    msg2 = MIMEMultipart('alternative')
    msg2['From'] = "%s <%s>" % ("", app.config['SMTP_SENDER'])
    msg2['To'] = 'ssh@kitech.re.kr'
#     msg2['To'] = 'akj2995@naver.com'
    msg2['Subject'] = "[생산기술연구원-인공지능특허검색시스템] 기술이전 정보 접수2"
    msg2.attach(MIMEText(message, 'html', 'utf-8'))  # 내용 인코딩
    server.sendmail(app.config['SMTP_SENDER'], msg2['To'], msg2.as_string())

manager.create_api(TechTrans
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='techtrans'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , postprocessors={'POST': [post_tech_trans]}
                   , preprocessors={
                        'PATCH_SINGLE': [check_token_single],
                        'GET_SINGLE': [check_token_single],
                        'GET_MANY': [check_token]
                   })

@app.route('/api/v1/upload_file', methods=['POST'])
def upload_file_api():
    print("1")
    check_token()
    print("2")
    print(request.files)
    if 'file' not in request.files:
        return make_response(jsonify({'result': False}), 400)

    f = request.files['file']
    if f.filename == '':
        return make_response(jsonify({'result': False}), 400)

    upload_path = os.getcwd() + '/uploads/'

    if f and '.' in f.filename:
        filename = f.filename
        f.save(os.path.join(upload_path, filename))
        return make_response(jsonify({'result': True, 'filename': filename,'src_filename':f.filename}), 200)


@app.route('/api/v1/attachment', methods=['POST'])
def set_attachment_images():
    print("1")
    check_token()
    print("2")
    print(request.files)
    if 'file' not in request.files:
        return make_response(jsonify({'result': False}), 400)

    f = request.files['file']
    if f.filename == '':
        return make_response(jsonify({'result': False}), 400)

    upload_path = os.getcwd() + '/attachments/'

    if f and '.' in f.filename:
        ext = f.filename.rsplit('.', 1)[1].lower()

        if ext in {'png', 'jpg', 'jpeg', 'gif'}:
            filename = str(uuid.uuid1()) + '.' + ext
            f.save(os.path.join(upload_path, filename))
            return make_response(jsonify({'result': True, 'filename': filename,'src_filename':f.filename}), 200)

@app.route('/api/v1/upload_file/<filename>', methods=['GET'])
def get_upload_fil_api(filename):
    upload_path = os.getcwd() + '/uploads/'
    return send_from_directory(upload_path, filename)

@app.route('/api/v1/attachment/<filename>', methods=['GET'])
def get_attachment_images(filename):
    upload_path = os.getcwd() + '/attachments/'
    return send_from_directory(upload_path, filename)

@app.route('/api/v1/download/<filename>', methods=['GET'])
def get_download_excel(filename):
    upload_path = os.getcwd() + '/download/excel/'
    return send_from_directory(upload_path, filename)

@app.route('/api/v1/thumbnail/<filename>', methods=['GET'])
def get_thumbnail_images(filename):
    upload_path = os.getcwd() + '/thumbnail/'
    return send_from_directory(upload_path, filename)

@app.route('/api/v1/pdf/<filename>', methods=['GET'])
def get_pdf_images(filename):
    upload_path = os.getcwd() + '/pdf/'
    return send_from_directory(upload_path, filename)

@app.route('/api/v1/download/<filename>', methods=['GET'])
def get_download_images(filename):
    upload_path = os.getcwd() + '/download/excel/'
    print("get_download_images :",upload_path)
    return send_from_directory(upload_path, filename)


@app.route('/api/v1/pdf_upload', methods=['POST'])
def pdf_upload_api():
    check_token()
    print(request.files)
    if 'file' not in request.files:
        return make_response(jsonify({'result': False}), 400)

    f = request.files['file']
    if f.filename == '':
        return make_response(jsonify({'result': False}), 400)

    upload_path = os.getcwd() + '/pdf/'

    if f and '.' in f.filename:
        ext = f.filename.rsplit('.', 1)[1].lower()

        if ext in {'pdf'}:
            filename = f.filename
            f.save(os.path.join(upload_path, filename))
            return make_response(jsonify({'result': True, 'filename': filename,'src_filename':f.filename}), 200)

@app.route('/api/v1/thumb_upload', methods=['POST'])
def thumb_upload_api():
    check_token()
    print(request.files)
    if 'file' not in request.files:
        return make_response(jsonify({'result': False}), 400)

    f = request.files['file']
    if f.filename == '':
        return make_response(jsonify({'result': False}), 400)

    upload_path = os.getcwd() + '/thumbnail/'
    print("upload_path :",upload_path)
    if f and '.' in f.filename:
        ext = f.filename.rsplit('.', 1)[1].lower()

        if ext in {'png', 'jpg', 'jpeg', 'gif'}:
            filename = f.filename
            f.save(os.path.join(upload_path, filename))
            return make_response(jsonify({'result': True, 'filename': filename,'src_filename':f.filename}), 200)

global download_index
download_index = 0
@app.route('/api/v1/patent_manage_download', methods=['get'])
def patent_manage_download_api():
    print("patent_manage_download_api start...")
    parser = reqparse.RequestParser()
    parser.add_argument("tech_type", type=str, location='args', required=True)
    parser.add_argument("years", type=str, location='args', required=True)
    parser.add_argument("search_string", type=str, location='args', required=True)
    args = parser.parse_args()
    tech_type = int(args['tech_type'])
    years = args['years']
    search_string = args['search_string']
    templete_path = './download/excel/templete/patent_manage_xl_template0.xlsx'

    #xl파일 형식지정
    center_align = Alignment(horizontal='center',
                            vertical='bottom',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=True,
                            indent=0)

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    filter_or_group = []
    filter_and_group = []
    if tech_type is not None and tech_type != -1 :
        filter_and_group.append(KitechSource.tech_type == tech_type)
    if years is not None and years != '전체' :
        filter_and_group.append(KitechSource.apply_date == years)
    if search_string is not None and search_string != '' :
        filter_or_group.append(KitechSource.person.like("%" + search_string + "%"))
        filter_or_group.append(KitechSource.patent_code.like("%" + search_string + "%"))
        filter_or_group.append(KitechSource.description.like("%" + search_string + "%"))

    source_data = KitechSource.query.filter(or_(*filter_or_group)).filter(and_(*filter_and_group)).order_by(KitechSource.id.desc()).all()
    print("len : ",len(source_data))
    wb = openpyxl.load_workbook(templete_path)
    ws = wb['Sheet1']
    xl_range = 7
    cur=0
    for rows in source_data:
        line_tuple = (
            rows.patent_code,
            rows.person,
            rows.apply_date,
            tech_types_dict[rows.tech_type],
            rows.invention_title[:30],
           rows.updated_date.strftime("%Y-%m-%d %H:%M:%S")
        )
        for i in range(xl_range):
            ws.cell(row=cur+1, column=i+1).border = thin_border
            ws.cell(row=cur+1, column=i+1).alignment = center_align
        cur+=1
        ws.append(line_tuple)

    for k in range(3):
        for i in range(xl_range):
            ws.cell(row=cur+1, column=i+1).border = thin_border
            ws.cell(row=cur+1, column=i+1).alignment = center_align
        cur+=1
    global download_index
    download_index += 1
    save_file_name = 'download_xl_file_{0}.xlsx'.format(download_index)
    save_file_path = './download/excel/' + save_file_name
    if os.path.isfile(save_file_path):
      os.remove(save_file_path)

    wb.save(save_file_path)
    res = {
        "filename":save_file_name
    }

#     print("type:",tab_item)
#     print(type(tab_item))
#     print("templete:",templete_path)
#     print('xl_range:',xl_range)
    return make_response(jsonify(res), 200)


@app.route('/api/v1/kitech_ai_search', methods=['GET'])
def kitech_ai_search_api():
    limit = request.args.get('limit')
    offset = request.args.get('offset')
    querystring = request.args.get('querystring')
    ai_use = request.args.get('ai_use')
    tech_type = request.args.get('tech_type')
    apply_date = request.args.get('apply_date')
    send_url = "http://localhost:6050/api/v1/search_ai?limit={0}&offset={1}&querystring={2}&ai_use={3}&tech_type={4}&apply_date={5}".format(limit, offset,querystring,ai_use,tech_type,apply_date)
    print("send_url : ",send_url)
    response = requests.get(send_url)
    print("response : ",response.text)
    return make_response(response.text, 200)
#
@app.route('/api/v1/addelasticsearch_kitech_fromdb', methods=['GET'])
def addelasticsearch_kitech_fromdb_api():
    print('addelasticsearch_kitech_fromdb ==> run')
    api_headers = {'Content-type': 'application/json'}

    def int_to_str(data):
        if data == None:
            text = 'None'
        elif type(data) == int:
            text = str(data)
        else :
            text = str(int(data))
        return text

    standard_data = TB_KITECH_PATENT_PARAGRAPH.query.all()
    cur = 1
    elastic_add_url = ELASTIC_SERVER + '/deepirext/_doc?pretty'
    for cur,data_line in enumerate(standard_data):
#         if cur < 7759 :
#             continue
        print("data_line.patent_code :",data_line.patent_code)
        src = KitechSource.query.filter_by(patent_code=data_line.patent_code).first()
        dataobj = {
            "doc_id":data_line.idx,
            "docs":data_line.patent_paragraph_expend.lower(),
            "patent_code":data_line.patent_code,
            "tech_type":str(src.tech_type).lower(),
            "years":str(src.apply_date).lower(),
        }
        print("dataobj :", dataobj)
        r = requests.post(elastic_add_url, verify=False, data=json.dumps(dataobj), headers=api_headers)
        print("r_text : ",json.loads(r.text)," / cur : ",str(cur + 1)+'\n')
        time.sleep(0.02)

    return make_response(jsonify({}), 200)

@app.route('/api/v1/addtechlist', methods=['GET'])
def addtechlist_api():
    from openpyxl import load_workbook
    print('addtechlist start')
    saved_xls_file = "./data/지식재산권현황.xlsx"
    wb = load_workbook(saved_xls_file, data_only=True)
    sheet = wb.active
    rows = sheet.rows

    cur = 0
    def getValue(value) :
        if value is None :
            return ''
        return value

    for row in rows:
        if cur <= 1 :
            cur +=1
            continue
        new_item = TechTypeList()
        new_item.main_person = str(row[0].value).strip()
        new_item.patent_title = row[1].value
        new_item.patent_code = str(row[2].value).replace('-','')
        new_item.tech_type_name = row[3].value
        db.session.add(new_item)

        cur +=1
    db.session.commit()
    return make_response(jsonify({}), 200)

@app.route('/api/v1/match_convert_techlist', methods=['GET'])
def match_convert_techlist_api():
    tech_list = TechTypeList.query.all()

    for tech in tech_list:
        if len(tech.patent_code) > 5 :
            patent = KitechSource.query.filter_by(patent_code=tech.patent_code).first()
            if patent is not None :
                if len(tech.tech_type_name) > 1 :
                    obj = {
                    "main_person":tech.main_person,
                    "tech_type":tech_types[tech.tech_type_name]
                    }
                    print("obj :",obj)
                    db.session.query(KitechSource).filter_by(patent_code=patent.patent_code).update(obj)
    db.session.commit()
    return make_response(jsonify({}), 200)

@app.route('/api/v1/statics_report', methods=['GET'])
def statics_report_api():
    parser = reqparse.RequestParser()
    args = parser.parse_args()
    response = dict()
    print("args :",args)

    q = KitechSource.query.with_entities(
        KitechSource.tech_type,
        func.count(KitechSource.id).label('count')) \
        .group_by(KitechSource.tech_type).all()

    response['tech_types_objs'] = [dict(tech_type=item.tech_type, count=item.count) for item in q]
    print("response['tech_types_objs'] :",response['tech_types_objs'] )

    q = KitechSource.query.with_entities(
        KitechSource.person,
        func.count(KitechSource.id).label('count')) \
        .group_by(KitechSource.person).order_by(func.count(KitechSource.id).desc()).limit(5).all()

    response['person_objs'] = [dict(person=item.person, count=item.count) for item in q]
    print("response['person_objs'] :",response['person_objs'] )

    q = KitechSource.query.with_entities(
        KitechSource.apply_date,
        func.count(KitechSource.id).label('count')) \
        .group_by(KitechSource.apply_date).order_by(KitechSource.apply_date.desc()).limit(5).all()

    response['apply_date_objs'] = [dict(apply_date=str(item.apply_date) + "년", count=item.count) for item in q]
    print("response['apply_date_objs'] :",response['apply_date_objs'] )

    q = KITECH_USER_ACTION.query.with_entities(
        KITECH_USER_ACTION.create_date,
        func.count(KITECH_USER_ACTION.id).label('count')) \
        .filter(KITECH_USER_ACTION.create_date >= (datetime.now()-timedelta(days=365))) \
        .group_by(func.year(KITECH_USER_ACTION.create_date),func.month(KITECH_USER_ACTION.create_date)).order_by(KITECH_USER_ACTION.create_date.asc()).limit(12).all()

    response['user_action_months_objs'] = [dict(month=item.create_date.strftime("%Y-%m"), count=item.count) for item in q]
    print("response['user_action_months_objs'] :",response['user_action_months_objs'] )

    q = KITECH_USER_ACTION.query.with_entities(
        KITECH_USER_ACTION.search_string,
        func.count(KITECH_USER_ACTION.id).label('count')) \
        .group_by(KITECH_USER_ACTION.patent_code).order_by(func.count(KITECH_USER_ACTION.id).desc()).limit(10).all()
    response['search_rank_objs'] = [dict(search_string=item.search_string, count=item.count) for item in q]
    print("response['search_rank_objs'] :",response['search_rank_objs'] )

    return make_response(jsonify(response), 200)

manager.create_api(KITECH_USER_ACTION
                   , results_per_page=10000
                   , url_prefix='/api/v1'
                   , collection_name='user_action'
                   , methods=['GET', 'DELETE', 'PATCH', 'POST']
                   , allow_patch_many=True
                   , preprocessors={
                   })